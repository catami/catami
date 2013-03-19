"""Helpers to import metadata (excel workbooks, csv etc.).
"""

__author__ = "Lachlan Toohey"

from openpyxl import load_workbook  # for xlsx
from openpyxl.cell import column_index_from_string

from xlrd import open_workbook, xldate_as_tuple
import os.path

from catamidb.models import BRUVDeployment, DOVDeployment, TIDeployment, TVDeployment

import datetime

import logging

logger = logging.getLogger(__name__)


def metadata_models():
    mapping = dict()
    mapping['bruv'] = BRUVDeployment
    mapping['dov'] = DOVDeployment
    mapping['ti'] = TIDeployment
    mapping['tv'] = TVDeployment

    return mapping


def xls_outline(metadata_file, heading_row=0):
    """Return a dictionary describing the file.

    The return is a dictionary with worksheet names as keys
    and values of lists with column headings as elements.

    This is designed for xls files.
    """

    workbook = open_workbook(metadata_file)
    sheets = workbook.sheet_names()
    descriptor = {}

    for sheet in sheets:
        worksheet = workbook.sheet_by_name(sheet)
        first_row = worksheet.row_values(heading_row)

        headings = []

        for val in first_row:
            headings.append(val)

        descriptor[sheet] = headings

    return descriptor


def xlsx_outline(metadata_file, heading_row=0):
    """Return a dictionary describing the file.

    The return is a dictionary with worksheet names as keys
    and values of lists with column headings as elements.

    This is designed for xlsx files.
    """

    workbook = load_workbook(metadata_file)
    sheets = workbook.get_sheet_names()
    descriptor = {}

    for sheet in sheets:
        worksheet = workbook.get_sheet_by_name(sheet)
        last_col = worksheet.get_highest_column()
        first_row = worksheet.rows[heading_row]

        headings = []

        for cell in first_row:
            headings.append(cell.value)

        descriptor[sheet] = headings

    return descriptor


def xls_transform(metadata_file, sheet_name, field_heading_map={},
                  shared_fields={}, heading_row=0):
    """Prepares rows of excel file to 'structure' format for further processing.

    Takes the excel file to use, the name of the worksheet to import from
    the mapping of fields to headings used in the model and any field values that
    are common across all rows (and optionally not specified).
    """
    logger.debug("xls_transform: opening workbook.")

    workbook = open_workbook(metadata_file)
    worksheet = workbook.sheet_by_name(sheet_name)

    # get the headings that are in the sheet
    heading_column_map = {}
    ncols = worksheet.row_len(heading_row)
    nrows = worksheet.nrows

    for col in xrange(ncols):
        heading_column_map[worksheet.cell_value(heading_row, col)] = col
    logger.debug(
        "xls_transform: found {0} headings.".format(len(heading_column_map)))

    # create a map from field to column directly
    # or use heading to column if there is a null mapping
    if field_heading_map == dict():
        field_column_map = heading_column_map
        logger.debug("xls_transform: using headings as fields.")
    else:
        field_column_map = {}
        logger.debug("xls_transform: mapping {0} fields.".format(
            len(field_heading_map)))
        for field, heading in field_heading_map.iteritems():
            field_column_map[field] = heading_column_map[heading]

    logger.debug(
        "xls_transform: mapping {0} columns.".format(len(field_column_map)))

    structure = []
    for row in xrange(heading_row + 1, nrows):
        # initialise using the shared base
        fields = dict(shared_fields)

        # get the values for each field
        for field, column in field_column_map.iteritems():
            cell = worksheet.cell(row, column)

            ctype = cell.ctype

            #if ftype == 1: # ie estimated DATE
            if ctype == 3:  # cell type marked as date
                try:
                    # if ambiguous throws an exception
                    output = xldate_as_tuple(cell.value, workbook.datemode)
                except:
                    # exception, just store raw value
                    fields[field] = cell.value
                else:
                    # got a tuple, process it!
                    if output[0] == 0:
                        # this is a time
                        fields[field] = datetime.time(*output[3:])
                    else:
                        # datetime!
                        # check that the time is not exactly midnight
                        # this is a heuristic to hopefully reduce incorrect
                        # recognitions of dates
                        if output[0] < 1950 and output[3] == 0 and output[
                            4] == 0 and output[5] == 0:
                            # time is midnight... year is too early, likely mis-recognised
                            fields[field] = cell.value
                        else:
                            fields[field] = datetime.datetime(*output)
            else:
                fields[field] = cell.value

        structure.append(fields)

    return structure


def xlsx_transform(metadata_file, sheet_name, field_heading_map={},
                   shared_fields={}, heading_row=0):
    """Prepares rows of excel file to 'structure' format for further processing.

    Takes the excel file to use, the name of the worksheet to import from
    the mapping of fields to headings used in the model and any field values that
    are common across all rows (and optionally not specified).
    """

    logger.debug("xlsx_transform: opening workbook.")

    workbook = load_workbook(metadata_file)
    worksheet = workbook.get_sheet_by_name(sheet_name)

    # get the headings that are in the sheet
    heading_column_map = {}
    for cell in worksheet.rows[heading_row]:
        # the index returned from fn is 1 based, later function needs 0 based
        heading_column_map[cell.value] = column_index_from_string(
            column=cell.column) - 1
    logger.debug(
        "xlsx_transform: found {0} headings.".format(len(heading_column_map)))

    # create a map from field to column directly
    # or use heading to column if there is a null mapping
    if field_heading_map == dict():
        field_column_map = heading_column_map
        logger.debug("xlsx_transform: using headings as fields.")
    else:
        field_column_map = {}
        logger.debug("xlsx_transform: mapping {0} fields.".format(
            len(field_heading_map)))
        for field, heading in field_heading_map.iteritems():
            field_column_map[field] = heading_column_map[heading]

    logger.debug(
        "xlsx_transform: mapping {0} columns.".format(len(field_column_map)))

    structure = []
    # now extract the real rows of interest
    for data_row in xrange(heading_row + 1, worksheet.get_highest_row() + 1):

        # initialise using the shared base
        fields = dict(shared_fields)

        # get the values for each field
        for field, column in field_column_map.iteritems():
            fields[field] = worksheet.cell(row=data_row, column=column).value

        structure.append(fields)

    return structure
